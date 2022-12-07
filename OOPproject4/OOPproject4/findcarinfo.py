import cv2
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import pytesseract
import os
from datetime import datetime
from find_chars import *
from isHangeul import *

#function that helps makes car object from picture
wb = openpyxl.load_workbook('car_info.xlsx')
w2 = wb['car_info']
w3 = wb['car_owner']

now = datetime.now()

def findcarinfo(image_name):

    #get date, time info from picture name(0000-00-00 00;00;00.jpg format)
    datetime = image_name.split('.')
    datesplittime = datetime[0].split()
    car_date = datesplittime[0]
    car_time = datesplittime[1]


    #matplotlib 이용해 이미지 파일 읽어 그래프로 나타냄
    img_ori = cv2.imread(image_name)
    #shape함수 이용해 height, width, channel 얻음
    height, width, channel = img_ori.shape
    plt.figure(figsize=(12, 10))
    plt.imshow(img_ori,cmap='gray')

    ###1st image, grayscale###
    gray = cv2.cvtColor(img_ori, cv2.COLOR_BGR2GRAY) #openCV 컬러변환 함수 cvtColor 이용해 흑백으로 변환
    plt.figure(figsize=(12,10)) #figure 인스턴스 생성, 이미지 전체의 영역 확보

    #cmap 옵션 이용해 변경할 colormap 입력, gray - 흑백옵션
    plt.imshow(gray, cmap='gray')

    ###2nd image, GaussianBlur and ThreshHolding###

    img_blurred = cv2.GaussianBlur(gray, ksize=(5, 5), sigmaX=0)

    #adaptiveThreshold(입력영상, 결과 이진 영상,결과에 대한 최댓값, 방법, 경계화 타입, 블록의 크기, 사용할 경계 값);
    #경계화 타입에 가우시안블러 적용, 중심 쪽 픽셀의 가중치 높게 부여
    img_blur_thresh = cv2.adaptiveThreshold(
        img_blurred,
        maxValue=255.0,
        adaptiveMethod=cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        thresholdType=cv2.THRESH_BINARY_INV,
        blockSize=19,
        C=9
    )

    ###3rd image, Finding Contour###
    #find contours: 동일한 색/강도 가진 영역의 경계선 연결한 선 찾기
    #LIST mode: 계층정보x, 모든 외곽선 검출/method:수직선, 수평선, 대각선 대해 끝점만 저장
    contours, _ = cv2.findContours(
        img_blur_thresh,
        mode=cv2.RETR_LIST,
        method=cv2.CHAIN_APPROX_SIMPLE
    )
    #0으로 초기화된 배열 생성, grayscale이미지
    temp_result = np.zeros((height, width, channel), dtype=np.uint8)
    #윤곽선 그리는 함수 drawContours
    cv2.drawContours(temp_result, contours=contours, contourIdx=-1, color=(255,255,255))
    plt.figure(figsize=(12, 10))
    plt.imshow(temp_result)

    ###4th image, Drawing rectangle on assumption number###
    temp_result = np.zeros((height, width, channel), dtype=np.uint8)

    #contours들의 좌표 이용해 사각형 형태로 그림
    contours_dict = []

    #boundingrect():주어진 점을 감싸는 최소 크기 사각형(바운딩 박스)를 반환
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        cv2.rectangle(temp_result, pt1=(x,y), pt2=(x+w, y+h), color=(255,255,255), thickness=2)
        
        #딕셔너리에 contours들의 정보 저장
        contours_dict.append({
            'contour': contour,
            'x': x,
            'y': y,
            'w': w,
            'h': h,
            'cx': x + (w / 2),
            'cy': y + (h / 2)
        })
        

    #번호판 글자인 것 같은 contours들 추려내기
    MIN_AREA = 80 # 번호판 윤곽선 최소 범위 지정
    MIN_WIDTH, MIN_HEIGHT=2, 8 # 최소 너비 높이 범위 지정
    MIN_RATIO, MAX_RATIO = 0.25, 1.0 # 최소 비율 범위 지정

    possible_contours = []

    cnt = 0
    for d in contours_dict:
        area = d['w'] * d['h']
        ratio = d['w'] / d['h']
        
        if area > MIN_AREA \
        and d['w'] > MIN_WIDTH and d['h'] > MIN_HEIGHT \
        and MIN_RATIO < ratio < MAX_RATIO:
            d['idx'] = cnt
            cnt += 1
            possible_contours.append(d)

    temp_result = np.zeros((height, width, channel), dtype = np.uint8)

    for d in possible_contours:
        #cv2.drawContours(temp_result, d['contour'], -1, (255, 255, 255))
        cv2.rectangle(temp_result, pt1=(d['x'], d['y']), pt2=(d['x']+d['w'], d['y']+d['h']), color=(255, 255, 255), thickness=2)

    plt.figure(figsize=(12,10))
    plt.imshow(temp_result, cmap='gray')


    ###6th image, Selecting Candidates (2)Contours###

    result_idx = find_chars(possible_contours)

    matched_result = []
    for idx_list in result_idx:
        matched_result.append(np.take(possible_contours, idx_list))

    # visualize possible contours
    temp_result = np.zeros((height, width, channel), dtype=np.uint8)

    for r in matched_result:
        for d in r:
    #         cv2.drawContours(temp_result, d['contour'], -1, (255, 255, 255))
            cv2.rectangle(temp_result, pt1=(d['x'], d['y']), pt2=(d['x']+d['w'], d['y']+d['h']), color=(255, 255, 255), thickness=2)

            

    plt.figure(figsize=(12, 10))
    plt.imshow(temp_result, cmap='gray')
    #현재까지 후보는 temp_result임

    PLATE_WIDTH_PADDING = 1.3 # 1.3
    PLATE_HEIGHT_PADDING = 1.5 # 1.5
    MIN_PLATE_RATIO = 3
    MAX_PLATE_RATIO = 10

    plate_imgs = []
    plate_infos = []

    for i, matched_chars in enumerate(matched_result):
        sorted_chars = sorted(matched_chars, key=lambda x: x['cx'])

        plate_cx = (sorted_chars[0]['cx'] + sorted_chars[-1]['cx']) / 2
        plate_cy = (sorted_chars[0]['cy'] + sorted_chars[-1]['cy']) / 2
        
        plate_width = (sorted_chars[-1]['x'] + sorted_chars[-1]['w'] - sorted_chars[0]['x']) * PLATE_WIDTH_PADDING
        
        sum_height = 0
        for d in sorted_chars:
            sum_height += d['h']

        plate_height = int(sum_height / len(sorted_chars) * PLATE_HEIGHT_PADDING)
        
        triangle_height = sorted_chars[-1]['cy'] - sorted_chars[0]['cy']
        triangle_hypotenus = np.linalg.norm(
            np.array([sorted_chars[0]['cx'], sorted_chars[0]['cy']]) - 
            np.array([sorted_chars[-1]['cx'], sorted_chars[-1]['cy']])
        )
        
        angle = np.degrees(np.arcsin(triangle_height / triangle_hypotenus))
        
        rotation_matrix = cv2.getRotationMatrix2D(center=(plate_cx, plate_cy), angle=angle, scale=1.0)
        
        img_rotated = cv2.warpAffine(img_blur_thresh, M=rotation_matrix, dsize=(width, height))
        
        img_cropped = cv2.getRectSubPix(
            img_rotated, 
            patchSize=(int(plate_width), int(plate_height)), 
            center=(int(plate_cx), int(plate_cy))
        )
        
        if img_cropped.shape[1] / img_cropped.shape[0] < MIN_PLATE_RATIO or img_cropped.shape[1] / img_cropped.shape[0] < MIN_PLATE_RATIO > MAX_PLATE_RATIO:
            continue
        
        plate_imgs.append(img_cropped)
        plate_infos.append({
            'x': int(plate_cx - plate_width / 2),
            'y': int(plate_cy - plate_height / 2),
            'w': int(plate_width),
            'h': int(plate_height)
        })
        
        plt.subplot(len(matched_result), 1, i+1)
        plt.imshow(img_cropped, cmap='gray')

    longest_idx, longest_text = -1, 0
    plate_chars = []

    for i, plate_img in enumerate(plate_imgs):
        plate_img = cv2.resize(plate_img, dsize=(0, 0), fx=1.6, fy=1.6)
        _, plate_img = cv2.threshold(plate_img, thresh=0.0, maxval=255.0, type=cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        
        # find contours again (same as above)
        contours, _ = cv2.findContours(plate_img, mode=cv2.RETR_LIST, method=cv2.CHAIN_APPROX_SIMPLE)
        
        plate_min_x, plate_min_y = plate_img.shape[1], plate_img.shape[0]
        plate_max_x, plate_max_y = 0, 0

        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            
            area = w * h
            ratio = w / h

            if area > MIN_AREA \
            and w > MIN_WIDTH and h > MIN_HEIGHT \
            and MIN_RATIO < ratio < MAX_RATIO:
                if x < plate_min_x:
                    plate_min_x = x
                if y < plate_min_y:
                    plate_min_y = y
                if x + w > plate_max_x:
                    plate_max_x = x + w
                if y + h > plate_max_y:
                    plate_max_y = y + h
                    
        img_result = plate_img[plate_min_y:plate_max_y, plate_min_x:plate_max_x]
         #왜곡 보정 후 가져온 이미지의 후보는  img_result
        
        ###9th image, Using tesseract  ###      
        img_result = cv2.GaussianBlur(img_result, ksize=(3, 3), sigmaX=0)
        _, img_result = cv2.threshold(img_result, thresh=0.0, maxval=255.0, type=cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        img_result = cv2.copyMakeBorder(img_result, top=10, bottom=10, left=10, right=10, borderType=cv2.BORDER_CONSTANT, value=(0,0,0))
        
        pytesseract.pytesseract.tesseract_cmd = 'C:\Program Files\Tesseract-OCR/tesseract.exe'
        chars = pytesseract.image_to_string(img_result, lang='kor', config='--psm 7 --oem 0')
        #chars에 일단 다 때려 박음
        
        result_chars = ''
        num_digit = 0
        num_hangeul = 0
        
        # [ 번호판의 조건 ]
        # 1. 숫자가 6개 또는 7개
        # 2. 한글은 1개
        # 3. 한글은 인덱스[2], [3]
        
         # 1. 숫자나 한글로만 범위를 한정
        for c in chars:
            if ord('가') <= ord(c) <= ord('힣') or c.isdigit():
                if c.isdigit(): #숫자 개수 확인
                    num_digit += 1
                elif isHangeul(c): #한글 개수 확인
                    num_hangeul +=1
                result_chars += c
                
        
        plate_chars.append(result_chars)
        # plate_chars 에는 후보인 번호판애들이 다 들어감!!!
        # result_chars 에는 번호판 하나씩 가지고 있음 (반복문 속 변수)
        #print(plate_chars)

        if num_digit and len(result_chars) > longest_text:
            longest_idx = i

        plt.subplot(len(plate_imgs), 1, i+1)
        plt.imshow(img_result, cmap='gray')

    info = plate_infos[longest_idx]
    chars = plate_chars[longest_idx]

    # 2. chars 에서 다시 검사 

    ###9th image, final###

    #번호판 정보를 텍스트로 콘솔에 출력

    if num_digit == 6 and num_hangeul == 1 :
        print(chars)
        row_count = wb["car_info"].max_row
        w2.cell(row=row_count + 1, column=1).value = str(now.strftime('%Y-%m-%d %H:%M:%S'))
        w2.cell(row=row_count + 1, column=2).value = chars

        row_count1 = wb["car_owner"].max_row
        for kk in range(2,row_count1+1):
            if chars == w3.cell(kk,2).value:
                w2.cell(row=row_count + 1, column=3).value = w3.cell(kk,4).value
                w2.cell(row=row_count + 1, column=4).value = w3.cell(kk, 5).value
                w2.cell(row=row_count + 1, column=5).value = w3.cell(kk, 6).value
    elif num_digit == 7 and num_hangeul == 1:
        print(chars)
        row_count = wb["car_info"].max_row
        w2.cell(row=row_count + 1, column=1).value = str(now.strftime('%Y-%m-%d %H:%M:%S'))
        w2.cell(row=row_count + 1, column=2).value = chars

        row_count1 = wb["car_owner"].max_row
        for kk in range(2,row_count1+1):
            if chars == w3.cell(kk,2).value:
                w2.cell(row=row_count + 1, column=3).value = w3.cell(kk,4).value
                w2.cell(row=row_count + 1, column=4).value = w3.cell(kk, 5).value
                w2.cell(row=row_count + 1, column=5).value = w3.cell(kk, 6).value
    else :
        print('다시 인식해주세요.')


    car_number = chars

    img_out = img_ori.copy()

    #rectangle(), 사각형을 그리는 함수
    cv2.rectangle(img_out, pt1=(info['x'], info['y']), pt2=(info['x']+info['w'], info['y']+info['h']), color=(255,0,0), thickness=2)

    new_img_name = chars + '.jpg'
    extension = os.path.splitext(new_img_name)[1]
    result, encoded_img = cv2.imencode(extension, img_out)
    if result:
        with open('res/'+new_img_name, mode='w+b') as f:
            encoded_img.tofile(f)


    plt.figure(figsize=(12, 10))
    plt.imshow(img_out)

    #객체 생성 위한 변수값 return
    #car = Car(car_date, car_time, car_number)
    wb.save('car_info.xlsx')
    wb.close()
    return car_date, car_time, car_number
