import cv2
import numpy as np
import matplotlib.pyplot as plt
import pytesseract
import os
import findcarinfo
import openpyxl
import pandas as pd
from tabulate import tabulate
plt.style.use('dark_background')

class Car:
    def __init__(self, cardate, cartime, carnum):
        self.__cardate = cardate
        self.__cartime = cartime
        self.__carnum = carnum

    def getcardate(self):
        return self.__cardate
    def getcartime(self):
        return self.__cartime
    def getcarnum(self):
        return self.__carnum
    def setcardate(self, newcardate):
        self.__carnum = newcardate
    def setcartime(self, newcartime):
        self.__carnum = newcartime
    def setcarnum(self, newcarnum):
        self.__carnum = newcarnum
    

#carlist saves car objects
carlist=[]

#function that makes car object from picture

    

#path:이미지 들어있는 경로    
#path = 'D:/2학년 2학기/객지프/프로젝트4'
path = 'C:/Users/newid/OneDrive/바탕 화면/workspace/python/OOP_proj4/OOPproject4'
img_list=[]
possible_img_extension = ['.jpg','.jpeg','.JPG','.bmp','.png', '.PNG']
for(root,dirs,files) in os.walk(path):
    dirs[:] = [dir for dir in dirs if dir!="res"]
    if len(files)>0:
        for file_name in files:
            if os.path.splitext(file_name)[1] in possible_img_extension:
                
                img_list.append(file_name)

for img in img_list:
    carname, cardate, cartime = findcarinfo.findcarinfo(img)
    car = Car(carname, cardate, cartime)
    carlist.append(car)

def find_owner_info_by_number(car_number, path):
    wb = openpyxl.load_workbook(path)
    ws1 = wb['car_info']
    ws2 = wb['car_owner']
    res = []
    for row in ws2:
        values = []
        for cell in row:
            values.append(cell.value)
        if values[1] == car_number:
            res.append(values)
    wb.close()
    df = pd.DataFrame(res, columns=['Datetime', 'Car Number', 'Car Type', 'Building Nr.', 'Apartment Nr.', 'Owner\'s name'])
    print(f'OWNER INFO BY CAR NUMBER {car_number}')
    print(tabulate(df, headers='keys', tablefmt='psql'))
    return values

def find_info_by_datetime(datetime, path):
    wb = openpyxl.load_workbook(path)
    ws1 = wb['car_info']
    cars_result = []
    for row in ws1:
        values = []
        for cell in row:
            values.append(cell.value)

        car_datetime = values[0]
        if car_datetime == datetime:
            cars_result.append(values)
    wb.close()
    df = pd.DataFrame(cars_result, columns=['Datetime', 'Car Number', 'Building Nr.', 'Apartment Nr.', 'Owner\'s name'])
    print(f'INFO BY DATETIME {datetime}')
    print(tabulate(df, headers='keys', tablefmt='psql'))
    return cars_result

def find_info_by_time(time, path):
    wb = openpyxl.load_workbook(path)
    ws1 = wb['car_info']
    cars_result = []
    for row in ws1:
        values = []
        for cell in row:
            values.append(cell.value)

        datetime = values[0]
        split_datetime = datetime.split(' ')
        if len(split_datetime) >= 2 and split_datetime[1] == time:
            cars_result.append(values)
    wb.close()
    df = pd.DataFrame(cars_result, columns=['Datetime', 'Car Number', 'Building Nr.', 'Apartment Nr.', 'Owner\'s name'])
    print(f'INFO BY TIME {time}')
    print(tabulate(df, headers='keys', tablefmt='psql'))
    return cars_result

def find_info_by_date(date, path):
    wb = openpyxl.load_workbook(path)
    ws1 = wb['car_info']
    cars_result = []
    for row in ws1:
        values = []
        for cell in row:
            values.append(cell.value)

        datetime = values[0]
        split_datetime = datetime.split(' ')
        if len(split_datetime) >= 2 and split_datetime[0] == date:
            cars_result.append(values)
    wb.close()
    print(f'INFO BY DATE {date}')
    df = pd.DataFrame(cars_result, columns=['Datetime', 'Car Number', 'Building Nr.', 'Apartment Nr.', 'Owner\'s name'])
    print(tabulate(df, headers='keys', tablefmt='psql'))
    return cars_result


# find_owner_info_by_number('65노0887', './car_info.xlsx')
# find_info_by_datetime('2022-12-05 21:46:43', './car_info.xlsx')
# find_info_by_date('2022-12-05', './car_info.xlsx')
# find_info_by_time('21:46:43', 'car_info.xlsx')

prompt = """1. Find owner info by car number
2. Find info by datetime
3. Find info by date
4. Find info by time
5. Exit"""

while True:
    os.system('cls')
    print (prompt)
    index = int(input("Enter number : "))
    if(index == 1):
        find = input("Enter car number : ")
        find_owner_info_by_number(find, './car_info.xlsx')
    elif(index == 2):
        find = input("Enter datetime : ")
        find_info_by_datetime(find, './car_info.xlsx')
    elif(index == 3):
        find = input("Enter date : ")
        find_info_by_date(find, './car_info.xlsx')
    elif(index == 4):
        find = input("Enter time : ")
        find_info_by_time(find, './car_info.xlsx')
    else:
        break
    input("Press enter to continue ")

#print(carlist[2].getcardate())
